import os
import xlrd
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

def get_metadata_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    props = wb.properties
    return {
        "Auteur": props.creator,
        "Dernière modification par": props.lastModifiedBy,
        "Création": props.created,
        "Modifié": props.modified
    }

def get_metadata_xls(file_path):
    wb = xlrd.open_workbook(file_path, on_demand=True)
    return {
        "Auteur": wb.user_name,
        "Date de création": xlrd.xldate_as_datetime(wb.datemode, 0) if wb.datemode else "Unknown"
    }

def get_file_metadata(file_path):
    file_info = {
        "Nom du fichier": os.path.basename(file_path),
        "Taille du fichier (KB)": round(os.path.getsize(file_path) / 1024, 2),
        "Dernière modification": datetime.fromtimestamp(os.path.getmtime(file_path)),
    }
    if file_path.endswith(".xlsx"):
        file_info.update(get_metadata_xlsx(file_path))
    elif file_path.endswith(".xls"):
        file_info.update(get_metadata_xls(file_path))
    else:
        raise ValueError("Format de fichier non supporté")
    return file_info

def get_columns_info(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    columns_info = {}
    for col_idx, col_cells in enumerate(ws.iter_cols(), start=1):
        column_name = col_cells[0].value if col_cells[0].value else f"Colonne {col_idx}"
        sample_value = next((cell.value for cell in col_cells[1:] if cell.value is not None), None)
        data_type = type(sample_value).__name__ if sample_value is not None else "Unknown"
        columns_info[column_name] = {"Type": data_type, "Définition": ""}
    return columns_info

def save_metadata_to_excel(metadata, columns_info, output_path, sujet_fichier):
    wb = Workbook()
    ws_meta = wb.active
    ws_meta.title = "Métadonnées"
    ws_meta.append(["Attribut", "Valeur"])
    for key, value in metadata.items():
        ws_meta.append([key, value])
    ws_meta.append(["Sujet des fichiers", sujet_fichier])
    ws_meta.append([])  
    ws_meta.append(["Nom de la colonne", "Type", "Définition"])
    for col, info in columns_info.items():
        ws_meta.append([col, info["Type"], info["Définition"]])
    wb.save(output_path)
    messagebox.showinfo("Succès", f"Fichier enregistré sous {output_path}")

def upload_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls;*.xlsx")])
    if not file_path:
        return
    metadata = get_file_metadata(file_path)
    columns_info = get_columns_info(file_path)
    
    def save():
        output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not output_path:
            return
        for col, entry in entries.items():
            columns_info[col]["Définition"] = entry.get()
        save_metadata_to_excel(metadata, columns_info, output_path, sujet_input.get())
    
    root = tk.Toplevel()
    root.title("Métadonnées du fichier")
    root.geometry("800x600")
    
    frame = ttk.Frame(root, padding=10)
    frame.pack(fill=tk.BOTH, expand=True)
    
    ttk.Label(frame, text="Métadonnées", font=("Arial", 14, "bold")).pack()
    for key, value in metadata.items():
        ttk.Label(frame, text=f"{key}: {value}", anchor="w", justify="left").pack(fill=tk.X)
    
    ttk.Label(frame, text="Sujet du fichier").pack()
    sujet_input = ttk.Entry(frame, width=50)
    sujet_input.pack()
    
    ttk.Separator(frame, orient="horizontal").pack(fill=tk.X, pady=10)
    ttk.Label(frame, text="Définition des attributs", font=("Arial", 12, "bold")).pack()
    columns_frame = ttk.Frame(frame)
    columns_frame.pack(fill=tk.BOTH, expand=True)
    
    canvas = tk.Canvas(columns_frame)
    scrollbar = ttk.Scrollbar(columns_frame, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas)
    
    scrollable_frame.bind(
        "<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")
        )
    )
    
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    entries = {}
    for col, info in columns_info.items():
        row_frame = ttk.Frame(scrollable_frame)
        row_frame.pack(fill=tk.X, pady=2)
        ttk.Label(row_frame, text=f"{col} ({info['Type']}):", anchor="w", justify="left").pack(side=tk.LEFT)
        entry = ttk.Entry(row_frame, width=40)
        entry.pack(side=tk.RIGHT)
        entries[col] = entry
    
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    ttk.Separator(frame, orient="horizontal").pack(fill=tk.X, pady=10)
    save_button = ttk.Button(frame, text="Enregistrer", command=save)
    save_button.pack()
    
    root.mainloop()

def main():
    root = tk.Tk()
    root.title("Extracteur de Métadonnées Excel")
    root.geometry("400x200")
    
    frame = ttk.Frame(root, padding=20)
    frame.pack(fill=tk.BOTH, expand=True)
    
    upload_button = ttk.Button(frame, text="Charger un fichier", command=upload_file)
    upload_button.pack()
    
    root.mainloop()

if __name__ == "__main__":
    main()
